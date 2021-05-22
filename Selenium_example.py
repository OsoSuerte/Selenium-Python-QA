from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
import selenium.webdriver.support.ui as ui
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import xlut
import logging
from pyzbar.pyzbar import decode
from PIL import Image
import cv2
import qrtools

'''
#Firefox set up
fp = webdriver.FirefoxProfile()
fp.set_preference('browser.helperApps.neverAsk.saveToDisk', 'text/plain, application/pdf')
fp.set_preference('browser.download.manager.showWhenStarting', False)
fp.set_preference('browser.download.dir', 'C:\Coding\Portfolio\Selenium\Voluteer Form')
fp.set_preference('browser.download.folderList', 2)
fp.set_preference('pdfjs.disabled', True)
driver = webdriver.Firefox(executable_path="C:\Drivers\geckodriver-v0.29.1-win64\geckodriver.exe", firefox_profile= fp)
'''

# Next three lines remove Chrome's "browser controlled by software" banner. 4th line changes dowload dir
options = webdriver.ChromeOptions()
options.add_experimental_option("useAutomationExtension", False)
options.add_experimental_option("excludeSwitches",  ["enable-automation"])
options.add_experimental_option('prefs', {'download.default_directory': 'C:\Coding\Portfolio\Selenium\Voluteer Form'})
driver = webdriver.Chrome(executable_path="C:\Drivers\chromedriver_win32\chromedriver.exe", options=options)  #Selects Chrome for webdriver

driver.get('https://testautomationpractice.blogspot.com')  #Selects chosen URL
driver.implicitly_wait(10)
driver.maximize_window()  #Maximizes browser window
action = ActionChains(driver)


#driver.switch_to.frame('frame-one1434677811')
#driver.switch_to_default_content()
#driver.switch_to.frame(0)  # Switches focus to indicated frame.

# Below sets and enables logging
logging.basicConfig(filename="C:\Coding\Portfolio\Selenium\Voluteer Form\logs\logtest.log", format='%(asctime)s: %(levelname)s: %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', level=logging.DEBUG)
logging.debug('This is debug')
logging.info('This is info')
logging.warning('This is the warning')
logging.error('This is the error')
logging.critical('This is critical')
driver.switch_to.default_content()


# Finds all links on the page with tag 'a' then iterates through them and returns the link text.
lnk = driver.find_elements_by_tag_name('a')
for i in lnk:
    if i.text ==  '':
        continue
    print('Link text is : ' + i.text)

print()
driver.switch_to.frame('frame-one1434677811')
input_boxes = driver.find_elements(By.CLASS_NAME, 'text_field')  # Counts the number of input boxes on form
boxes = driver.find_elements(By.CLASS_NAME, 'multiple_choice')  ## Counts the number of radios on form
print('There are ' + str(len(input_boxes)) + ' input boxes on the page')  # Prints the number of input boxes on form.
print('There are ' + str(len(boxes)) + ' radios on the page')  # Prints the number of radios on form.

#This for loop is just an example of how to iterate through objects and show a status, displayed, selected, etc.
count = 1
print()
for i in input_boxes:
    print('Input box is displayed = ' + str(driver.find_element(By.ID, 'RESULT_TextField-' + str(count)).is_displayed()))
    print('Input box is enabled = ' + str(driver.find_element(By.ID, 'RESULT_TextField-' + str(count)).is_enabled()))
    print('***')
    count += 1

driver.switch_to.default_content()
# Below sends text to wiki search box and selects the first result, opens in new tab than closes tab
driver.find_element_by_id("Wikipedia1_wikipedia-search-input").send_keys('selenium')
driver.find_element_by_xpath('//*[@id="Wikipedia1_wikipedia-search-form"]/div/span[2]/span[2]/input').click()
driver.find_element_by_xpath('//*[@id="wikipedia-search-result-link"]/a').click()
driver.switch_to.window(driver.window_handles[1])
time.sleep(1)
driver.close()
driver.switch_to.window(driver.window_handles[0])  # Changes back to parent window

# Below click a button that creates an alert and then accepts/dismisses the alert
driver.find_element_by_xpath('//*[@id="HTML9"]/div[1]/button').click()
time.sleep(.5)
driver.switch_to.alert.accept()
driver.find_element_by_xpath('//*[@id="HTML9"]/div[1]/button').click()
time.sleep(.5)
driver.switch_to.alert.dismiss()

driver.find_element_by_id('datepicker').send_keys('04/07/2024')  # Sends date to date picker

Select(driver.find_element_by_xpath('//*[@id="speed"]')).select_by_visible_text('Faster')
Select(driver.find_element_by_xpath('//*[@id="files"]')).select_by_index(3)
Select(driver.find_element_by_xpath('//*[@id="number"]')).select_by_visible_text('3')
Select(driver.find_element_by_xpath('//*[@id="products"]')).select_by_value('Apple')
Select(driver.find_element_by_xpath('//*[@id="animals"]')).select_by_index(1)


def bid(id, text):  # Function for a quick  find by ID.
    return driver.find_element_by_id(id).send_keys(text)
def sel(id):  # Function for a quick click by id.
    return driver.find_element_by_id(id).click()
driver.switch_to.frame('frame-one1434677811')  # Switch to indicated frame
#Next six lines fill text boxes with the associated text
bid('RESULT_TextField-1', 'Oso')
bid('RESULT_TextField-2', 'Suerte')
bid('RESULT_TextField-3', '555-555-5555')
bid('RESULT_TextField-4', 'USA')
bid('RESULT_TextField-5', 'America Town')
bid('RESULT_TextField-6', 'foo@bar.com')

driver.find_element_by_tag_name('html').send_keys(Keys.PAGE_DOWN)  # Pages down to bring new elements into window
driver.find_element_by_css_selector("label[for='RESULT_RadioButton-7_0']").click()  # Clicks male radio
time.sleep(.5)
driver.find_element_by_css_selector("label[for='RESULT_RadioButton-7_1']").click()  #Clicks female radio
time.sleep(.5)

#  For loop to iterate through and select each check box on the form.
count = 0
for i in range(7):
    time.sleep(.3)
    driver.find_element_by_css_selector("label[for='RESULT_CheckBox-8_" + str(count) + "']").click()
    print('Box was marked ' + str(driver.find_element_by_id('RESULT_CheckBox-8_' + str(count)).is_selected()))  # verifes click
    print('***')  # Used to show break between iterations
    count += 1
# Below selects dropdown box option using different methods.
drp = Select(driver.find_element_by_id("RESULT_RadioButton-9"))
drp.select_by_visible_text('Morning')
time.sleep(.5)
drp.select_by_index(2)
time.sleep(.5)
drp.select_by_value('Radio-2')

#Prints number of dropdown options and their text
print('There are ' + str(len(drp.options)) + ' dropdown options' )
print('The options are:')
for option in drp.options:
    print(option.text)

driver.switch_to.default_content()  #Switches back to default content from current frame
driver.find_element_by_xpath('//*[@id="field1"]').clear()  # Removes text from double click field 1
driver.find_element_by_xpath('//*[@id="field1"]').send_keys('Big badda boom!')  # sends text to above field
dbclk = driver.find_element_by_xpath('//*[@id="HTML10"]/div[1]/button')  # Sets "copy text" button to variable
ActionChains(driver).double_click(dbclk).perform()  # Sends double click command to variable
source = driver.find_element_by_xpath('//*[@id="draggable"]/p')
target = driver.find_element_by_xpath('//*[@id="droppable"]')
ActionChains(driver).drag_and_drop(source, target).perform()  # Drags source to target

#below sends male and female image to trash ussing drag and drop
source = driver.find_element_by_xpath('//*[@id="gallery"]/li/img')
target = driver.find_element_by_xpath('//*[@id="trash"]')
ActionChains(driver).drag_and_drop(source, target).perform()
time.sleep(.5)
source = driver.find_element_by_xpath('//*[@id="gallery"]/li/img')
#target = driver.find_element_by_xpath('//*[@id="trash"]')
ActionChains(driver).drag_and_drop(source, target).perform()
#ActionChains(driver).click_and_hold(source).move_to_element(target).release(target).perform()

# Below moves slider and resizes text box.
slide = driver.find_element_by_id("slider")
ActionChains(driver).click_and_hold(slide).move_by_offset(50, 0).release().perform()
driver.execute_script("window.scrollBy(0, 800)", "")
res = driver.find_element_by_xpath('//*[@id="resizable"]/div[3]')
ActionChains(driver).click_and_hold(res).move_by_offset(250, 200).release().perform()

age = driver.find_element_by_xpath('//*[@id="age"]')  # Sets "your age" box to variable.
age.send_keys('100')  # Sends text to box
ActionChains(driver).move_to_element(age).perform()  # clicks varaible hovered over
print()
#  Below download qr png file read, print data in QR code
url = 'https://2.bp.blogspot.com/-pvlE2uQU6Jg/XCx2wNhWs8I/AAAAAAAAPPs/upCq39nWswU839wODhr6xtimoukMzEPBQCLcBGAs/s1600/qrcode.png'
urllib.request.urlretrieve(url, "Test.png")
img = cv2.imread('C:\Coding\Portfolio\Selenium\Voluteer Form\Test.png')
for i in decode(img):
    print('The QR code says ' + i.data.decode('utf-8') + "'")
#img2 = decode(img)  # alternate way to print QR data
#print(img2[0][0].decode('utf-8'))

# below is working with tables, first find the number of rows and columns
driver.switch_to.default_content()  # Switches from frame to default

rows = len(driver.find_elements_by_xpath('//*[@id="HTML1"]/div[1]/table/tbody/tr'))
print()
print('There are ' + str(rows) + ' rows in the table')
col = len(driver.find_elements_by_xpath('//*[@id="HTML1"]/div[1]/table/tbody/tr[1]/th'))
print('There are ' + str(col) + ' columns in the table')
#iterate through the table and print its data
print()
print('The below data can be found in the table:')

#iterate through the table and print its data


for r in range(2, rows + 1):
    for c in range(1, col + 1):
       val = driver.find_element_by_xpath('//*[@id="HTML1"]/div[1]/table/tbody/tr[' + str(r) + ']/td[' + str(c) + ']').text
       print(val, end = '    ')
    print()

driver.switch_to.frame('frame-one1434677811')  # Switches focus to indicated frame.
# Open the link in a new tab by sending key strokes on the element
# Use: Keys.CONTROL + Keys.SHIFT + Keys.RETURN to open tab on top of the stack
driver.find_element_by_link_text('Software Testing Tutorials').send_keys(Keys.CONTROL + Keys.RETURN)
time.sleep(1)
driver.find_element_by_link_text('Software Testing Tools Training').send_keys(Keys.CONTROL + Keys.RETURN)
time.sleep(1)
driver.find_element_by_xpath('//*[@id="FSForm"]/div[2]/div[27]/a[2]').send_keys(Keys.CONTROL + Keys.RETURN)
time.sleep(1)
driver.find_element_by_xpath('//*[@id="FSForm"]/div[2]/div[27]/a[1]').send_keys(Keys.CONTROL + Keys.RETURN)
time.sleep(1)
# Switch to open tabs by index: will populate in reverse of the order they were opened.
driver.switch_to.window(driver.window_handles[4])
time.sleep(1)
driver.switch_to.window(driver.window_handles[3])
time.sleep(1)
driver.switch_to.window(driver.window_handles[2])
time.sleep(1)
driver.switch_to.window(driver.window_handles[1])
time.sleep(1)
# Close the tabs we opened one at a time.

for i in driver.window_handles:
    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])
        driver.close()
        time.sleep(.5)
driver.switch_to.window(driver.window_handles[0])

driver.close()

#*********** un-used WORKING CODE ********

'''
#  Below sets vairbale and right clicks on the variable.
rclk = driver.find_element_by_xpath('//*[@id="HTML4"]/div[1]/img')
action.context_click(rclk).perform()

# Below are two ways to get a screen shot
driver.save_screenshot(r"C:\Coding\Portfolio\Selenium\Voluteer Form\seltest.jpeg")
driver.get_screenshot_as_file(r'C:\Coding\Portfolio\Selenium\Voluteer Form\selte.png')

# Below working with cookies
co = {'name': 'Mycookie', 'value': 'ososuerte'}  # Create new cookie as dictioanry
driver.add_cookie(co)  # Will add coockie to page
driver.delete_cookie(co)  # Will delete cookie from page 
driver.delete_all_cookies()  #Deletes all cookies on a page
cook = driver.get_cookies()  #Saves all cookies to a variable
print(len(cook))
print(cook)

# Below will open xcell file and read/write save data.
path = 'C:\Coding\Portfolio\Selenium\Voluteer Form\ss.xlsx'  # Path of file
workbook = openpyxl.load_workbook(path) 
sheet = workbook.active  # Selects active sheet
#sheet = workbook.get_sheet_by_name('Sheet1')  # Selects sheet my name
rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)
for r in range(1, 88):
    sheet.cell(row = r, column = 3).value = 'new'
for r in range(1, 88):
    sheet.cell(row=r, column=2).value = 'TAX'
for r in range(1, 88):
    for c in range(1, 4):
        print(sheet.cell(row = r, column = c).value,end = ' - ')
    print()
workbook.save(path)
#  Below can be used for functions of the above.
#  You can also save data to a veriable and then send it to a page text box.
#path = 'C:\Coding\Portfolio\Selenium\Voluteer Form\ss.xlsx'  # Path of file
#rownum = xlut.rowct(path, 'Sheet1')
#for r in range(1, rows + 1):

# Different ways to scroll
ele = driver.find_element_by_xpath('//*[@id="HTML4"]/div[1]/img')  
driver.execute_script('arguments[0].scrollIntoView();', ele)  # Scrolls to the element given as argument
driver.execute_script('window.scrollBy(0,document.body.scrollHeight)')  #Scrolls to bottom of page. 
driver.execute_script("window.scrollBy(0, 1800)", "")  # scroll down use positive numbers
driver.execute_script("window.scrollBy(0, -1800)", "")  # scroll up use negative numbers

#driver.find_element_by_id('RESULT_FileUpload-10').send_keys(r'C:\Coding\Portfolio\Selenium\Voluteer Form\#upload.txt')  # Will upload file to target. 

#  Below will download file based on url. 
url = 'https://2.bp.blogspot.com/-dw3qYanXdhM/Wg6dD9-99RI/AAAAAAAAMso/SSJ2vCg-f9QkXneDP3kp6AJPv4EeU1aSwCLcBGAs/s1600/male.png'
urllib.request.urlretrieve(url, "Test.png")  # 2nd par will be name of file when downloaded


# Below clicks a link navigates back and then selects a link using a different method
driver.find_element_by_link_text('Software Testing Tutorials').click()
time.sleep(1)
driver.back()
#time.sleep(3)
driver.find_element_by_partial_link_text('Training').click()

#below are other ways that work to check the radio. 
action = ActionChains(driver)
element = driver.find_element_by_id('RESULT_RadioButton-7_0')
action.move_to_element(element).click().perform()
driver.find_element_by_xpath('//*[@id="q15"]/table/tbody/tr[1]/td/label').click() #checks box 8_0
btn = driver.find_element_by_id('RESULT_RadioButton-7_0')
driver.execute_script ("arguments[0].click();",btn)
'''


